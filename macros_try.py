import openpyxl
from datetime import datetime, timedelta

wb = openpyxl.load_workbook("prac.xlsx")
sheets_to_process = ["5000UA501", "5120ZI415IOP", "5000UA062", "5660XS634", "all"]  # список листов для обработки

tag_data = {}
intervals = {}

for sheet_name in sheets_to_process:
    ws = wb[sheet_name]

    header = {cell.value.lower(): idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    time_col = header.get("time")
    status_col = header.get("status")
    tag_col = header.get("tag")

    if not all([time_col, status_col, tag_col]):
        raise ValueError(f"Не удалось найти колонки 'time', 'status' и 'tag' в таблице {sheet_name}.")

    data = [
        (row[time_col-1], row[status_col-1], row[tag_col-1])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)
    ]

    if not data:
        raise ValueError(f"Лист {sheet_name} не содержит данных.")

    # обработка данных по минутным интервалам
    for time_val, status, tag in data:
        current_time = datetime.combine(datetime.min, time_val)
        interval_key = (tag, current_time - timedelta(seconds=current_time.second, microseconds=current_time.microsecond))

        if interval_key not in intervals:
            intervals[interval_key] = {
                "alm_count": 0,
                "suppressing": False,
                "suppressed_count": 0
            }

        interval = intervals[interval_key]

        if status == "ALM":
            interval["alm_count"] += 1

            if interval["suppressing"]:
                interval["suppressed_count"] += 1
            else:
                if interval["alm_count"] > 3:
                    interval["suppressing"] = True
                    interval["suppressed_count"] = interval["alm_count"] - 3

        if tag not in tag_data:
            tag_data[tag] = {
                "total_alm_count": 0,
                "total_suppressed_signals": 0
            }

        tag_data[tag]["total_alm_count"] += 1

    # подсчёт подавленных сигналов для каждого тэга в каждом интервале
    for (tag, interval_start), interval in intervals.items():
        tag_data[tag]["total_suppressed_signals"] += interval["suppressed_count"]

result_ws = wb.create_sheet("Result")

result_ws.append(["Tag", "Total ALM Count", "Total Suppressed Signals"])

# записываем данные для каждого тэга
for tag, counts in tag_data.items():
    result_ws.append([tag, counts["total_alm_count"], counts["total_suppressed_signals"]])

wb.save("try2.xlsx")



























